import marimo

__generated_with = "0.14.10"
app = marimo.App(width="medium")


@app.cell
def _(mo):
    mo.md("""## **Company Valuation using Comparable Analysis** \n ***""")
    return


@app.cell(hide_code=True)
def _():
    import marimo as mo
    import pandas as pd
    #import polars as pl
    import numpy as np
    import altair as alt
    from sklearn.linear_model import LinearRegression
    import openpyxl
    import tabulate
    return LinearRegression, alt, mo, np, pd


@app.cell(hide_code=True)
def _(np, pd):
    # Read raw excel file
    DataRaw = pd.read_csv("Data/Data_ASEANFirms.csv",dtype=str)
    Data = DataRaw.iloc[:,:]
    # Convert string to float from column 8 onwards
    allColnames = DataRaw.columns # all column names
    cols_to_convert = allColnames[8:]
    #
    for c in cols_to_convert:
        Data[c] = pd.to_numeric(DataRaw[c],errors='coerce')
    #
    #Get unique value of primary industry and industry
    PrimaryIndustry = np.sort(Data["Primary Industry"].dropna().unique())
    Industry = np.sort(Data["Industry"].dropna().unique())
    # create dictionary with industry as key and sub-industry as value
    IndustryDict ={}
    for ind in Industry:
        IndustryDict [ind] = Data["Primary Industry"][Data["Industry"]==ind].unique()
    #
    CamGDP = 45150.0 # in USD'M
    return CamGDP, Data, IndustryDict


@app.cell
def _(IndustryDict, mo):
    # Create downdown ui for the industry
    Indus_dd = mo.ui.dropdown(options=IndustryDict,
                              label="Select Industry", 
                              value="Consumer Finance",searchable=True)
    return (Indus_dd,)


@app.cell
def _(Indus_dd):
    Indus_dd # Industry dropdown ui
    return


@app.cell
def _(Indus_dd, mo):
    # create checkboxes of subindustries once an industry is selected
    checkboxes = mo.ui.array([mo.ui.checkbox(label=subin, value=True) for subin in Indus_dd.value])
    return (checkboxes,)


@app.cell
def _(mo):
    mo.md(r"""**Choose one or more sub-industries**""")
    return


@app.cell
def _(checkboxes, mo):
    mo.vstack(checkboxes)
    return


@app.cell
def _(Data, Indus_dd, checkboxes, pd):
    subInd = pd.DataFrame({"SubInd": Indus_dd.value, "Selected":checkboxes.value}) # convert pd Dataframe
    selectedSubIndustry = subInd["SubInd"][subInd["Selected"]==True].to_numpy()
    usedData = Data[Data["Primary Industry"].isin(selectedSubIndustry)].copy()

    if Indus_dd.selected_key in ["Banks", "Consumer Finance", "Financial Services", "Insurance"]:
        bankFinancial = True
    else:
        bankFinancial = False
    #
    return bankFinancial, usedData


@app.cell
def _(mo):
    keywords_ui = mo.ui.text(placeholder="Search in business description....", label="Keywords:", full_width=True)
    keywords_ui
    return (keywords_ui,)


@app.cell
def _(keywords_ui, np, usedData):
    keywords = keywords_ui.value
    kws = keywords.split(";")    
    #
    if keywords=="":
        matchedKW = [True for elem in usedData["Business Description"]]
    else:
        matchedKW = [sum([elem.find(kw)>0 for kw in kws])>0 for elem in usedData["Business Description"]]
    #
    usedData["matchedKW"] = np.array(matchedKW)
    filteredData_1 = usedData.loc[usedData["matchedKW"],:].iloc[:,0:-1]
    return filteredData_1, matchedKW


@app.cell
def _(matchedKW, mo):
    if matchedKW.count(True)>=5:
        message1 = "**"+str(matchedKW.count(True)) + "** *firms matched!*"
    else:
        message1 = "**"+str(matchedKW.count(True)) + "** *firms matched! Try to add more key words*"
    mo.md(message1)
    return


@app.cell
def _(mo):
    prob_z_score_L = {"1.0%":-2.3263,"2.5%":-1.96, "5.0%": -1.645, "10%": -1.28}
    prob_z_score_U = {"99.0%":2.3263,"97.5%":1.96, "95.0%": 1.645, "90%": 1.28}
    L_dd = mo.ui.dropdown(options=prob_z_score_L, label="Lower Bound", value="2.5%")
    U_dd = mo.ui.dropdown(options=prob_z_score_U, label="Upper Bound", value="97.5%")
    return L_dd, U_dd


@app.cell
def _(L_dd, U_dd, mo):
    mo.vstack([mo.md("***"), mo.md("**Set outlier:** "),L_dd, U_dd,mo.md("***")])
    return


@app.cell
def _(bankFinancial, mo):
    # Target company inputs
    # Target company data for plot
    if bankFinancial:
        requiredInputs = ["Net Income (US$M)", "Total Equity(US$M)", "ROE (%)"]
        optionalInputs = ["Company Name", "Total Asset(US$M)", "Asset Rank"]
    else:
        requiredInputs = ["EBITDA (US$M)", "Total Revenue(US$M)", "Cash (US$M)","Debt (US$'M)"]
        optionalInputs = ["Company Name","Net Income (US$M)", "D/E"]
    #
    reqInputs_mo = mo.ui.array([mo.ui.text(placeholder="required...", label=input) for input in requiredInputs])
    optInputs_mo = mo.ui.array([mo.ui.text(placeholder="optional...", label=input) for input in optionalInputs])
    #
    mo.vstack([mo.md("**Provide Target Company Info:**")
        ,mo.hstack([mo.vstack(reqInputs_mo), mo.vstack(optInputs_mo)])])
    return optInputs_mo, reqInputs_mo


@app.cell
def _(CamGDP, bankFinancial, np, optInputs_mo, pd, reqInputs_mo):
    # collect inputs
    if bankFinancial:

        if optInputs_mo.value[0]=="":
            comName = "Target"
        else:
            comName = optInputs_mo.value[0]
        #
        if optInputs_mo.value[2]=="":
            assetRank = np.nan
        else:
            assetRank = int(optInputs_mo.value[2])
        if optInputs_mo.value[1]=="":
            total_asset = np.nan
            asset_GDP = np.nan
        else:
            total_asset = float(optInputs_mo.value[1])
            asset_GDP = total_asset*100.0/CamGDP
        #
        reqInputs = [float(i) if i!="" else np.nan for i in reqInputs_mo.value]
        #
        target_=pd.DataFrame({'Company Name':[comName], 'Net Income':[reqInputs[0]], 'Total Equity, LTM': [reqInputs[1]], 'ROE, LTM':[reqInputs[2]], 'RankByAssets': [assetRank], "Total Assset": [total_asset], 'Asset-to-GDP': [asset_GDP]})
    else:
        #print(bankFinancial)
        if optInputs_mo.value[0]=="":
            comName = "Target"
        else:
            comName = optInputs_mo.value[0]
        #
        if optInputs_mo.value[1]=="":
            NI = np.nan
        else:
            NI = float(optInputs_mo.value[1])
        if optInputs_mo.value[2]=="":
            DE = np.nan
        else:
            DE = float(optInputs_mo.value[2])
        #
        reqInputs = [float(i) if i!="" else np.nan for i in reqInputs_mo.value]
        #
        target_=pd.DataFrame({'Company Name':[comName], 'Net Income':[NI], 'EBITDA_IFRS, LTM': [reqInputs[0]], 'Total Revenue, LTM':[reqInputs[1]], 'EBITDA Margin, LTM':[100.0*reqInputs[0]/reqInputs[1]], "Revenue-to-GDP": [reqInputs[0]*100/CamGDP], "D/E": [DE], "Cash": reqInputs[2], "Debt": reqInputs[3]})
    #target_
    return (target_,)


@app.cell
def _(np):
    def filterBy(peerDF, targetDF, keyVar, properties):
        """
        Filter peer by target data to find peer similar to the target
        Inputs:
            - peerDF: (dataframe) peer data
            - targetDF: (dataframe) target data
            - keyVar: (string) key info of the target based on which the data will be filter
            - properties: a dictionary
                  maxiter : (int) max iteration
                  minnMatched: (int) min number of matched
                  step: step 
        Outputs: 
            - [simPeers, nMatched]
               simPeers: A Series of boolean (True; False) specifying which row will be selected
               nMatched: number of matched
        """
        #
        tarVar = targetDF[keyVar][0]
        maxiter = properties["maxiter"]
        minnMatched = properties["minnMatched"]
        step = properties["step"]
        #
        if np.isnan(tarVar)==False:
            iter = 0
            #

            simPeers = (peerDF[keyVar]>=tarVar-step) & (peerDF[keyVar]<=tarVar+step) #similar peers 
            nMatched = simPeers.tolist().count(True) # number of matched similar peers

            while (iter<maxiter and nMatched<minnMatched):
                iter +=1 
                simPeers = (peerDF[keyVar]>=tarVar-step*iter) & (peerDF[keyVar]<=tarVar+step*iter)
                nMatched = simPeers.tolist().count(True)
                #
                #print(f"iter: {iter}, nMatched: {nMatched}")

        else:
            simPeers = (peerDF[keyVar]!=np.nan)
            nMatched = simPeers.tolist().count(True)
        #print(f"iter: {iter}, nMatched: {nMatched}")
        return simPeers
    return (filterBy,)


@app.cell
def _(bankFinancial, filterBy, filteredData_1, target_):
    if bankFinancial:
        properties = {}
        properties["maxiter"] = 20
        properties["minnMatched"] = 15
        properties["step"] = filteredData_1["Asset-to-GDP"].std()*0.1
        sim_assetGDP = filterBy(filteredData_1,target_,"Asset-to-GDP",properties)
        #
        properties['step'] =1
        sim_assetRank = filterBy(filteredData_1,target_,"RankByAssets",properties)
        #
        simPeers = sim_assetGDP & sim_assetRank
    else:
        properties = {}
        properties["maxiter"] = 30
        properties["minnMatched"] = 15
        properties["step"] = filteredData_1["Revenue-to-GDP"].std()*0.1
        sim_revenueGDP = filterBy(filteredData_1,target_,"Revenue-to-GDP",properties)
        #
        properties["step"] = filteredData_1["D/E"].std()*0.1
        sim_DE = filterBy(filteredData_1,target_,"D/E",properties)
        #
        properties["step"] = filteredData_1["EBITDA Margin, LTM"].std()*0.1
        sim_EBITDAMargin = filterBy(filteredData_1,target_,"EBITDA Margin, LTM",properties)
        simPeers = sim_revenueGDP & sim_DE & sim_EBITDAMargin
    #
    filteredData = filteredData_1[simPeers]
    nMatched = simPeers.tolist().count(True)
    return filteredData, nMatched


@app.cell
def _(mo, nMatched):
    msg2 = "**" + str(nMatched) + "** *matched by Similarity*"
    mo.md(msg2)
    return


@app.cell
def _(L_dd, U_dd, filteredData):
    #
    # Find outliers of filtered data from column 8 onwards
    LBound = L_dd.value
    UBound = U_dd.value
    #
    nrows = filteredData.shape[0]
    std =filteredData.iloc[:,8:].std()
    mu = filteredData.iloc[:,8:].mean()
    normFilteredData = filteredData.iloc[:,8:].copy()
    for cn in filteredData.iloc[:,8:].columns:
        normFilteredData[cn] = (filteredData[cn]-mu[cn])/std[cn]
    # dataframe that store True/False for all key metrics
    Ex_Outliers = (normFilteredData>LBound) & (normFilteredData<UBound)
    # Treated outliers by setting them none
    treatedData = filteredData.copy()
    for j in Ex_Outliers.columns:
        treatedData.loc[~Ex_Outliers[j],j]=None
    return Ex_Outliers, treatedData


@app.cell
def _(LinearRegression, bankFinancial, nMatched, np, target_, treatedData):
    # Valuation 
    if bankFinancial:
        # Based on P/BV, LTM
        medianPB = treatedData['P/BV, LTM'].median()
        if ~np.isnan(target_["Total Equity, LTM"])[0]:
            PB_Valuation = medianPB*target_["Total Equity, LTM"]
        else:
            PB_Valuation = np.nan
        # Based on P/E Normalized, LTM
        medianPE = treatedData['P/E, Normalized, LTM'].median()
        if ~np.isnan(target_["Net Income"])[0]:
            PE_Valuation = medianPE*target_["Net Income"]
        else:
            PE_Valuation = np.nan
        # Based on regression model
        model = LinearRegression()
        Xname = "ROE, LTM"
        Yname = "P/BV, LTM"
        xy = treatedData[["Company Name",Xname, Yname]].dropna() # drop nulls
        model.fit(xy[[Xname]], xy[Yname])
        R2 = model.score(xy[[Xname]], xy[Yname])
        #
        if (R2>=0.25 and nMatched>=5 and ~np.isnan(target_[Xname])[0]):
            target_[Yname]=model.predict(target_[[Xname]])
        else:
            target_[Yname] = np.nan
        Reg_Valuation = target_[Yname]*target_["Total Equity, LTM"]
        #
        target_['P/BV_Valuation'] = round(PB_Valuation,1)
        target_['P/E_Valuation'] = round(PE_Valuation,1)
        target_['P/BV-ROE_Valuation'] = round(Reg_Valuation,1)
        ValuationMethods = ['P/BV_Valuation','P/E_Valuation','P/BV-ROE_Valuation']
    else:
        # Based on EV/EBITDA, LTM
        medianEVEBITDA = treatedData['TEV/EBITDA, LTM'].median()
        if ~np.isnan(target_["EBITDA_IFRS, LTM"])[0]:
            EVEBITDA_Valuation = medianEVEBITDA*target_["EBITDA_IFRS, LTM"]-target_["Debt"]+target_["Cash"]
        else:
            EVEBITDA_Valuation = np.nan
        # Based on P/E Normalized, LTM
        medianPE = treatedData['P/E, Normalized, LTM'].median()
        if ~np.isnan(target_["Net Income"])[0]:
            PE_Valuation = medianPE*target_["Net Income"]
        else:
            PE_Valuation = np.nan
        # Based on regression model
        model = LinearRegression()
        Xname = "EBITDA Margin, LTM"
        Yname = "TEV/Total Revenue, LTM"
        xy = treatedData[["Company Name",Xname, Yname]].dropna() # drop nulls
        model.fit(xy[[Xname]], xy[Yname])
        R2 = model.score(xy[[Xname]], xy[Yname])
        #
        if (R2>=0.25 and nMatched>=5 and ~np.isnan(target_[Xname])[0]):
            target_[Yname]=model.predict(target_[[Xname]])
        else:
            target_[Yname] = np.nan
        Reg_Valuation = target_[Yname]*target_["Total Revenue, LTM"] -target_["Debt"]+target_["Cash"]
        #
        target_['EV/EBITDA_Valuation'] = round(EVEBITDA_Valuation,1)
        target_['P/E_Valuation'] = round(PE_Valuation,1)
        target_['TEV/Revenue-EBITDAMargin_Valuation'] = round(Reg_Valuation,1)
        ValuationMethods = ['EV/EBITDA_Valuation','P/E_Valuation','TEV/Revenue-EBITDAMargin_Valuation']

    return (ValuationMethods,)


@app.cell
def _(ValuationMethods, alt, mo, pd, target_, treatedData):
    # Melt the DataFrame to long format for easier plotting
    target_long = pd.melt(target_, value_vars=ValuationMethods, var_name='Valuation_Method', value_name='Valuation')

    # Create the horizontal bar chart with data labels
    ValRes_BarChart = alt.Chart(target_long).mark_bar(size=30).encode(
        y=alt.Y('Valuation_Method:N', title='Valuation Method'),
        x=alt.X('Valuation:Q', title='Valuation'),
        tooltip=['Valuation_Method', alt.Tooltip('Valuation', format=",.1f")]
    ).properties(
        title='Company Valuation',
        width='container',
        height = 200
    )
    ValRes_BarChart_text = ValRes_BarChart.mark_text(
        align='left',
        baseline='middle',
        color ='red',
        dx=3  # Adjust horizontal offset
    ).encode(
        text=alt.Text('Valuation:Q', format=",.1f")
    )

    Valuation = target_[ValuationMethods].copy()
    subtitle1 = mo.md("###**Company Valuation:**")
    priceDate = mo.md("*Based on stock price of peers as of "+treatedData.columns[8][11:]+"*")
    mo.vstack([mo.md("***"),subtitle1, priceDate,ValRes_BarChart + ValRes_BarChart_text, Valuation])
    return


@app.cell
def _(mo):
    mo.md(
        r"""
    ***
    ###**Histogram and descriptive statitics**
    """
    )
    return


@app.cell
def _(bankFinancial, filteredData, mo):
    # Create downdown ui for columns
    if bankFinancial:
        default_metric = "P/BV, LTM"
    else:
        default_metric = "TEV/EBITDA, LTM"

    Metrics_dd = mo.ui.dropdown(options=filteredData.columns[8:],
                              label="Select Key Metrics 1", 
                              value=default_metric)
    Metrics_dd2 = mo.ui.dropdown(options=filteredData.columns[8:],
                              label="Select Key Metrics 2", 
                              value="P/E, Normalized, LTM")

    mo.accordion({"Show metrics👉":mo.hstack([Metrics_dd, Metrics_dd2])})
    return Metrics_dd, Metrics_dd2


@app.cell
def _(Ex_Outliers, Hist_keymetrics, Metrics_dd, filteredData, mo, pd):
    selMetric = filteredData[["No",Metrics_dd.value]]
    selMetric_exOut = selMetric[Ex_Outliers[Metrics_dd.value]==True] # excluded outliers
    #
    Chart_1 = Hist_keymetrics(selMetric_exOut)
    #
    desc1_pd = pd.DataFrame(selMetric_exOut.iloc[:, 1].describe())
    # use statistic as index and transpose
    desc1_pd = desc1_pd.T
    # change number format
    for col in desc1_pd.columns:
        desc1_pd[col] = desc1_pd[col].apply(lambda x: f"{x:.2f}")

    table1 = mo.md(desc1_pd.to_markdown(index=False))
    return Chart_1, table1


@app.cell
def _(Ex_Outliers, Hist_keymetrics, Metrics_dd2, filteredData, mo, pd):
    selMetric2 = filteredData[["No",Metrics_dd2.value]]
    selMetric2_exOut = selMetric2[Ex_Outliers[Metrics_dd2.value]==True] # excluded outliers
    #
    Chart_2 = Hist_keymetrics(selMetric2_exOut)
    #
    desc2_pd = pd.DataFrame(selMetric2_exOut.iloc[:, 1].describe())
    # use statistic as index and transpose
    desc2_pd = desc2_pd.T
    # change number format
    for col2 in desc2_pd.columns:
        desc2_pd[col2] = desc2_pd[col2].apply(lambda x: f"{x:.2f}")

    table2 = mo.md(desc2_pd.to_markdown(index=False))
    return Chart_2, table2


@app.cell
def _(alt):
    def Hist_keymetrics(keyMetric):
        hisChart = (
            alt.Chart(keyMetric)
            .mark_bar()
            .encode(
                x=alt.X(keyMetric.columns[1], type="quantitative", bin=True, title=keyMetric.columns[1]),
                y=alt.Y("count()", type="quantitative", title="Number of records"),
                tooltip=[
                    alt.Tooltip(
                        keyMetric.columns[1],
                        type="quantitative",
                        bin=True,
                        title=keyMetric.columns[1],
                        format=",.2f",
                    ),
                    alt.Tooltip(
                        "count()",
                        type="quantitative",
                        format=",.0f",
                        title="Number of records",
                    ),
                ],
            ).properties(width='container',height=250)
        )
        return hisChart

    return (Hist_keymetrics,)


@app.cell
def _(Chart_1, Chart_2, mo, table1, table2):
    histogramDiv = mo.accordion({"Show histogram 👉": mo.hstack([mo.vstack([Chart_1,table1]),mo.vstack([Chart_2,table2])])})
    histogramDiv
    return


@app.cell(hide_code=True)
def _(LinearRegression, alt):
    def scatterplot(peerData,TargetData, Xname, Yname):
        heig = 400
        xy = peerData[["Company Name",Xname, Yname]].dropna() # drop nulls
        #-----------------Base Scatter Plot---------------
        scatter = alt.Chart(xy).mark_point(filled=True,size =50.0).encode(
                x=alt.X(field=Xname, type='quantitative'),
                y=alt.Y(field=Yname, type='quantitative'),
                #color='Peer_Target',
                tooltip=[
                    alt.Tooltip(field="Company Name"),
                    alt.Tooltip(field=Xname, format=',.2f'),
                    alt.Tooltip(field=Yname, format=',.2f')
                ]
            ).properties(
                title=Xname + " vs " + Yname + " Scatter Plot with Regression Line",
                height=heig,
                width='container'
        )

        # Add equation line and text
        # Regression
        model = LinearRegression()
        # If data is more than 2 points
        if xy.shape[0]>1:
            model.fit(xy[[Xname]], xy[Yname])
            R2 = model.score(xy[[Xname]], xy[Yname])    
            ##--------------
            eqtext = f'y = {model.coef_[0]:.2f} x + {model.intercept_:.2f};' +f'\n(R2 = {R2:.2f})'
            #
            reg_line = scatter.transform_regression(
                Xname, Yname,
                method="linear",
            ).mark_line()

            scatterText = reg_line.mark_text(
                align='left',
                baseline='top',
                fontSize=13,
                dx=10,
                dy=10,
                text=eqtext,
                color = 'cyan'
            ).encode(
                x=alt.value(10),  # Position text
                y=alt.value(10)
                )
            ##---------
            if ~TargetData[Xname].isnull()[0]:
                # Target company data for plot
                TData_extended = TargetData.copy() 
                TData_extended[Yname]=model.predict(TargetData[[Xname]])
                #
                # add label
                textLabel =f"{TData_extended['Company Name'][0]} ({TData_extended[Xname][0]:.2f}, {TData_extended[Yname][0]:.2f})"

                TData_extended['label'] = textLabel
                #-----------------------------------------------------------------
                scatter_tar = alt.Chart(TData_extended).mark_point(filled=True,size =100.0,color="red").encode(
                        x=alt.X(field=Xname, type='quantitative'),
                        y=alt.Y(field=Yname, type='quantitative'),
                        #color='Peer_Target',
                        tooltip=[
                            alt.Tooltip(field="Company Name"),
                            alt.Tooltip(field=Xname, format=',.2f'),
                            alt.Tooltip(field=Yname, format=',.2f')
                        ]
                    ).properties(
                        title = Xname + " vs " + Yname + " Scatter Plot with Regression Line",
                        height = heig,
                        width='container',
                    )
                scatter_tar_text = alt.Chart(TData_extended).mark_text(
                    dx = -15, 
                    dy = -15,
                    color = "red",
                    fontSize= 15,
                    ).encode(
                        x=alt.X(field=Xname, type='quantitative'),
                        y=alt.Y(field=Yname, type='quantitative'),
                        text = 'label'
                )
                chart = scatter + scatterText + reg_line + scatter_tar + scatter_tar_text
            else:
                chart = scatter + scatterText + reg_line
        else:
            chart = scatter
        return chart
    return (scatterplot,)


@app.cell
def _(bankFinancial, mo, scatterplot, target_, treatedData):
    if bankFinancial:
        target=target_[["Company Name", "ROE, LTM"]]
        chart3 = scatterplot(treatedData, target,"ROE, LTM", "P/BV, LTM")
    else:
        target=target_[["Company Name", "EBITDA Margin, LTM"]]
        chart3 = scatterplot(treatedData, target,"EBITDA Margin, LTM", "TEV/Total Revenue, LTM")


    subtitle_div = mo.md("###**ScatterPlot with Regression Line**")
    scatterplotdiv = mo.accordion({'Scatter Plot with regression👉': chart3})
    mo.vstack([subtitle_div, scatterplotdiv])
    return


@app.cell
def _(mo, treatedData):
    sub_div2 = mo.md("###**Table of ASEAN peers**")
    tablediv = mo.accordion({"Table of matched ASEAN firms👉": treatedData})

    mo.vstack([sub_div2,tablediv])
    return


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
